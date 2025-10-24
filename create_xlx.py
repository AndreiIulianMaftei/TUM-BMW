import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

def load_json_data(json_file):
    """Load JSON data from file"""
    with open(json_file, 'r') as f:
        return json.load(f)

def format_currency(value):
    """Format number as currency"""
    if value is None:
        return "N/A"
    if value >= 1_000_000_000:
        return f"${value / 1_000_000_000:.1f}B"
    elif value >= 1_000_000:
        return f"${value / 1_000_000:.1f}M"
    elif value >= 1_000:
        return f"${value / 1_000:.1f}K"
    else:
        return f"${value:.2f}"

def create_summary_sheet(wb, data):
    """Create a summary overview sheet with all metrics"""
    ws = wb.active
    ws.title = "Executive Summary"
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_font = Font(bold=True, size=11, color="FFFFFF")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "Business Metrics Dashboard"
    ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Date
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Generated on: {datetime.now().strftime('%B %d, %Y')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(italic=True, size=10)
    
    # Headers
    row = 4
    headers = ['Metric Category', 'Key Value', 'Confidence', 'Insight']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row = 5
    for category, metrics in data.items():
        # Category name
        ws.cell(row=row, column=1, value=category).font = Font(bold=True, size=11)
        
        # Determine key value
        key_value = None
        if category == "TAM":
            key_value = format_currency(metrics.get("Market size"))
        elif category == "SAM":
            key_value = format_currency(metrics.get("Market size"))
        elif category == "SOM":
            key_value = format_currency(metrics.get("Revenue Potential"))
        elif category == "ROI":
            roi = metrics.get("ROI (%)")
            key_value = f"{roi}%" if roi else "N/A"
        elif category == "Turnover":
            key_value = format_currency(metrics.get("Total Revenue"))
        elif category == "Volume":
            units = metrics.get("Units Sold")
            key_value = f"{units:,}" if units else "N/A"
        elif category == "Unit Economics":
            margin = metrics.get("Margin")
            key_value = f"${margin}" if margin else "N/A"
        elif category == "EBIT":
            key_value = format_currency(metrics.get("EBIT Margin"))
        elif category == "COGS":
            key_value = format_currency(metrics.get("Total COGS"))
        elif category == "Market Potential":
            key_value = "90/100"
        
        ws.cell(row=row, column=2, value=key_value)
        
        # Confidence
        confidence = metrics.get("Confidence (%)")
        conf_cell = ws.cell(row=row, column=3, value=f"{confidence}%" if confidence else "N/A")
        
        # Color code confidence
        if confidence:
            if confidence >= 80:
                conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif confidence >= 70:
                conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Insight
        ws.cell(row=row, column=4, value=metrics.get("Insight", ""))
        
        # Apply borders
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='center')
        
        row += 1
    
    # Auto-adjust row heights
    for row_num in range(5, row):
        ws.row_dimensions[row_num].height = 40

def create_market_analysis_sheet(wb, data):
    """Create sheet with market analysis (TAM, SAM, SOM)"""
    ws = wb.create_sheet("Market Analysis")
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "Market Size Analysis (TAM-SAM-SOM)"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    row = 3
    ws['A3'] = "Market Level"
    ws['B3'] = "Market Size"
    ws['C3'] = "Confidence"
    
    for col in range(1, 4):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Data
    tam_value = data["TAM"]["Market size"]
    sam_value = data["SAM"]["Market size"]
    som_value = data["SOM"]["Revenue Potential"]
    
    ws['A4'] = "TAM (Total Addressable Market)"
    ws['B4'] = tam_value
    ws['C4'] = f"{data['TAM']['Confidence (%)']}%"
    
    ws['A5'] = "SAM (Serviceable Available Market)"
    ws['B5'] = sam_value
    ws['C5'] = f"{data['SAM']['Confidence (%)']}%"
    
    ws['A6'] = "SOM (Serviceable Obtainable Market)"
    ws['B6'] = som_value
    ws['C6'] = f"{data['SOM']['Confidence (%)']}%"
    
    # Create funnel chart
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Market Funnel: TAM → SAM → SOM"
    chart.y_axis.title = "Market Level"
    chart.x_axis.title = "Market Size ($)"
    
    data_ref = Reference(ws, min_col=2, min_row=3, max_row=6, max_col=2)
    cats = Reference(ws, min_col=1, min_row=4, max_row=6)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    
    ws.add_chart(chart, "E3")
    
    # Add insights
    ws['A10'] = "Key Insights:"
    ws['A10'].font = Font(bold=True, size=12)
    
    ws['A11'] = f"• {data['TAM']['Insight']}"
    ws['A12'] = f"• {data['SAM']['Insight']}"
    ws['A13'] = f"• {data['SOM']['Insight']}"
    
    # Calculate penetration rates
    sam_penetration = (sam_value / tam_value * 100) if tam_value else 0
    som_penetration = (som_value / sam_value * 100) if sam_value else 0
    
    ws['A15'] = "Penetration Analysis:"
    ws['A15'].font = Font(bold=True, size=12)
    ws['A16'] = f"• SAM represents {sam_penetration:.1f}% of TAM"
    ws['A17'] = f"• SOM represents {som_penetration:.1f}% of SAM"

def create_financial_metrics_sheet(wb, data):
    """Create sheet with financial metrics"""
    ws = wb.create_sheet("Financial Metrics")
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "Financial Performance Metrics"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Key metrics table
    row = 3
    metrics_data = [
        ["Metric", "Value", "Confidence"],
        ["Total Revenue (Turnover)", data["Turnover"]["Total Revenue"], f"{data['Turnover']['Confidence (%)']}%"],
        ["EBIT", data["EBIT"]["EBIT Margin"], f"{data['EBIT']['Confidence (%)']}%"],
        ["COGS", data["COGS"]["Total COGS"], f"{data['COGS']['Confidence (%)']}%"],
        ["ROI", f"{data['ROI']['ROI (%)']}%", f"{data['ROI']['Confidence (%)']}%"],
        ["Unit Margin", f"${data['Unit Economics']['Margin']}", f"{data['Unit Economics']['Confidence (%)']}%"],
    ]
    
    for r_idx, row_data in enumerate(metrics_data, start=row):
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == row:  # Header row
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
    
    # Create pie chart for revenue breakdown
    ws['E3'] = "Revenue"
    ws['F3'] = "Amount"
    ws['E4'] = "EBIT"
    ws['F4'] = data["EBIT"]["EBIT Margin"]
    ws['E5'] = "COGS"
    ws['F5'] = data["COGS"]["Total COGS"]
    ws['E6'] = "Gross Profit"
    ws['F6'] = data["Turnover"]["Total Revenue"] - data["COGS"]["Total COGS"]
    
    pie = PieChart()
    labels = Reference(ws, min_col=5, min_row=4, max_row=6)
    data_ref = Reference(ws, min_col=6, min_row=3, max_row=6)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Revenue Distribution"
    pie.height = 10
    pie.width = 15
    
    ws.add_chart(pie, "E10")
    
    # Add insights
    ws['A12'] = "Key Insights:"
    ws['A12'].font = Font(bold=True, size=12)
    ws['A13'] = f"• {data['Turnover']['Insight']}"
    ws['A14'] = f"• {data['EBIT']['Insight']}"
    ws['A15'] = f"• {data['COGS']['Insight']}"
    ws['A16'] = f"• {data['ROI']['Insight']}"

def create_confidence_analysis_sheet(wb, data):
    """Create sheet analyzing confidence levels"""
    ws = wb.create_sheet("Confidence Analysis")
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "Confidence Level Analysis"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    ws['A3'] = "Metric"
    ws['B3'] = "Confidence (%)"
    
    for col in [1, 2]:
        cell = ws.cell(row=3, column=col)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Data
    row = 4
    confidence_data = []
    for category, metrics in data.items():
        confidence = metrics.get("Confidence (%)")
        if confidence:
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=confidence)
            confidence_data.append((category, confidence))
            
            # Color code
            conf_cell = ws.cell(row=row, column=2)
            if confidence >= 80:
                conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif confidence >= 70:
                conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
    
    # Create bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Confidence Levels by Metric"
    chart.y_axis.title = "Confidence (%)"
    chart.x_axis.title = "Metric"
    
    data_ref = Reference(ws, min_col=2, min_row=3, max_row=row-1)
    cats = Reference(ws, min_col=1, min_row=4, max_row=row-1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 20
    
    ws.add_chart(chart, "D3")
    
    # Statistics
    avg_confidence = sum(c for _, c in confidence_data) / len(confidence_data)
    ws[f'A{row+2}'] = "Average Confidence:"
    ws[f'B{row+2}'] = f"{avg_confidence:.1f}%"
    ws[f'B{row+2}'].font = Font(bold=True)

def create_unit_economics_sheet(wb, data):
    """Create sheet for unit economics and volume"""
    ws = wb.create_sheet("Unit Economics")
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "Unit Economics & Volume Analysis"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Unit Economics
    ws['A3'] = "Unit Economics:"
    ws['A3'].font = Font(bold=True, size=12)
    
    ws['A4'] = "Profit Margin per Unit"
    ws['B4'] = f"${data['Unit Economics']['Margin']}"
    
    ws['A5'] = "Confidence"
    ws['B5'] = f"{data['Unit Economics']['Confidence (%)']}%"
    
    # Volume
    ws['A7'] = "Volume Metrics:"
    ws['A7'].font = Font(bold=True, size=12)
    
    ws['A8'] = "Units Sold (Annual)"
    ws['B8'] = data['Volume']['Units Sold']
    
    ws['A9'] = "Confidence"
    ws['B9'] = f"{data['Volume']['Confidence (%)']}%"
    
    # Calculate total profit from units
    total_profit = data['Unit Economics']['Margin'] * data['Volume']['Units Sold']
    ws['A11'] = "Projected Total Profit:"
    ws['A11'].font = Font(bold=True, size=12)
    ws['B11'] = format_currency(total_profit)
    ws['B11'].font = Font(bold=True, size=12, color="006100")
    
    # Insights
    ws['A13'] = "Key Insights:"
    ws['A13'].font = Font(bold=True, size=12)
    ws['A14'] = f"• {data['Unit Economics']['Insight']}"
    ws['A15'] = f"• {data['Volume']['Insight']}"

def main():
    """Main function to create the Excel file"""
    # Load data
    json_file = 'formatted_output.json'
    data = load_json_data(json_file)
    
    # Create workbook
    wb = Workbook()
    
    # Create all sheets
    print("Creating Executive Summary...")
    create_summary_sheet(wb, data)
    
    print("Creating Market Analysis...")
    create_market_analysis_sheet(wb, data)
    
    print("Creating Financial Metrics...")
    create_financial_metrics_sheet(wb, data)
    
    print("Creating Confidence Analysis...")
    create_confidence_analysis_sheet(wb, data)
    
    print("Creating Unit Economics...")
    create_unit_economics_sheet(wb, data)
    
    # Save workbook
    output_file = 'xlsx_files/business_metrics_dashboard.xlsx'
    wb.save(output_file)
    print(f"\n✅ Excel file created successfully: {output_file}")
    print(f"📊 Includes 5 sheets with comprehensive analysis and visualizations")

if __name__ == "__main__":
    main()
