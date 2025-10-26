"""
Professional Excel Export Module
Generates comprehensive, beautifully formatted Excel reports for BMW market analysis
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, Any, List
import os


class ExcelExporter:
    """Handles creation of professional Excel exports"""
    
    def __init__(self):
        # Define professional styling
        self.header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.subheader_font = Font(color="FFFFFF", bold=True, size=11)
        self.title_font = Font(bold=True, size=14, color="1F4788")
        self.normal_font = Font(size=10)
        self.currency_font = Font(size=10, bold=True)
        self.border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
    
    def style_header_row(self, ws, row, columns):
        """Apply header styling to a row"""
        for col in range(1, columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
    
    def style_subheader_row(self, ws, row, columns):
        """Apply subheader styling to a row"""
        for col in range(1, columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.subheader_fill
            cell.font = self.subheader_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def format_currency(self, value, currency="EUR"):
        """Format currency values"""
        if value is None:
            return "N/A"
        if isinstance(value, (int, float)):
            return f"€{value:,.2f}" if currency == "EUR" else f"${value:,.2f}"
        return str(value)
    
    def create_executive_summary_sheet(self, wb: Workbook, analysis_data: Dict[str, Any]):
        """Create Executive Summary sheet"""
        ws = wb.active
        ws.title = "Executive Summary"
        
        row = 1
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "BMW Market Analysis - Executive Summary"
        ws[f'A{row}'].font = Font(bold=True, size=16, color="1F4788")
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}"
        ws[f'A{row}'].font = Font(size=10, italic=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        # Project Overview
        if 'project_name' in analysis_data:
            ws[f'A{row}'] = "Project Name:"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = analysis_data.get('project_name', 'N/A')
            row += 1
        
        if 'executive_summary' in analysis_data:
            row += 1
            ws[f'A{row}'] = "Executive Summary"
            ws[f'A{row}'].font = self.title_font
            row += 1
            ws.merge_cells(f'A{row}:E{row+3}')
            ws[f'A{row}'] = analysis_data.get('executive_summary', '')
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.row_dimensions[row].height = 80
            row += 5
        
        # Market Potential Summary Table
        row += 1
        ws[f'A{row}'] = "Market Potential Summary"
        ws[f'A{row}'].font = self.title_font
        row += 1
        
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'C{row}'] = "Description"
        ws[f'D{row}'] = "Confidence"
        self.style_header_row(ws, row, 4)
        row += 1
        
        # TAM
        tam = analysis_data.get('tam', {})
        ws[f'A{row}'] = "TAM (Total Addressable Market)"
        ws[f'B{row}'] = self.format_currency(tam.get('market_size'))
        ws[f'C{row}'] = tam.get('description_of_public', 'N/A')
        ws[f'D{row}'] = f"{tam.get('confidence', 0)}%"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # SAM
        sam = analysis_data.get('sam', {})
        ws[f'A{row}'] = "SAM (Serviceable Addressable Market)"
        ws[f'B{row}'] = self.format_currency(sam.get('market_size'))
        ws[f'C{row}'] = sam.get('description_of_public', 'N/A')
        ws[f'D{row}'] = f"{sam.get('confidence', 0)}%"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # SOM
        som = analysis_data.get('som', {})
        ws[f'A{row}'] = "SOM (Serviceable Obtainable Market)"
        ws[f'B{row}'] = self.format_currency(som.get('revenue_potential'))
        ws[f'C{row}'] = som.get('description_of_public', 'N/A')
        ws[f'D{row}'] = f"{som.get('confidence', 0)}%"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = self.border
        row += 2
        
        # Financial Metrics
        ws[f'A{row}'] = "Key Financial Metrics"
        ws[f'A{row}'].font = self.title_font
        row += 1
        
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'C{row}'] = "Insight"
        self.style_header_row(ws, row, 3)
        row += 1
        
        # ROI
        roi = analysis_data.get('roi', {})
        ws[f'A{row}'] = "ROI"
        ws[f'B{row}'] = f"{roi.get('roi_percentage', 0):.1f}%"
        ws[f'C{row}'] = roi.get('insight', 'N/A')
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # Unit Economics
        unit_econ = analysis_data.get('unit_economics', {})
        ws[f'A{row}'] = "Unit Margin"
        ws[f'B{row}'] = f"{unit_econ.get('margin_percentage', 0):.1f}%"
        ws[f'C{row}'] = unit_econ.get('insight', 'N/A')
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # EBIT
        ebit = analysis_data.get('ebit', {})
        ws[f'A{row}'] = "EBIT Margin"
        ws[f'B{row}'] = f"{ebit.get('ebit_percentage', 0):.1f}%"
        ws[f'C{row}'] = ebit.get('insight', 'N/A')
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = self.border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 12
    
    def create_cost_breakdown_sheet(self, wb: Workbook, analysis_data: Dict[str, Any]):
        """Create Cost Breakdown sheet"""
        ws = wb.create_sheet("Cost Breakdown")
        row = 1
        
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Comprehensive Cost Analysis"
        ws[f'A{row}'].font = Font(bold=True, size=16, color="1F4788")
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        # Development Costs
        dev_costs = analysis_data.get('development_costs', [])
        if dev_costs:
            ws[f'A{row}'] = "Development Costs"
            ws[f'A{row}'].font = self.title_font
            row += 1
            
            ws[f'A{row}'] = "Category"
            ws[f'B{row}'] = "Amount (€)"
            ws[f'C{row}'] = "Reasoning"
            self.style_header_row(ws, row, 3)
            row += 1
            
            for cost in dev_costs:
                ws[f'A{row}'] = cost.get('category', 'N/A')
                ws[f'B{row}'] = cost.get('estimated_amount', 0)
                ws[f'B{row}'].number_format = '€#,##0.00'
                ws[f'C{row}'] = cost.get('reasoning', 'N/A')
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = self.border
                row += 1
            
            total_dev = analysis_data.get('total_development_cost', 0)
            ws[f'A{row}'] = "Total Development Cost"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = total_dev
            ws[f'B{row}'].number_format = '€#,##0.00'
            ws[f'B{row}'].font = self.currency_font
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = self.border
            row += 2
        
        # Customer Acquisition Costs
        cac_costs = analysis_data.get('customer_acquisition_costs', [])
        if cac_costs:
            ws[f'A{row}'] = "Customer Acquisition Costs"
            ws[f'A{row}'].font = self.title_font
            row += 1
            
            ws[f'A{row}'] = "Category"
            ws[f'B{row}'] = "Annual Budget (€)"
            ws[f'C{row}'] = "Reasoning"
            self.style_header_row(ws, row, 3)
            row += 1
            
            for cost in cac_costs:
                ws[f'A{row}'] = cost.get('category', 'N/A')
                ws[f'B{row}'] = cost.get('estimated_annual_budget', 0)
                ws[f'B{row}'].number_format = '€#,##0.00'
                ws[f'C{row}'] = cost.get('reasoning', 'N/A')
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = self.border
                row += 1
            row += 1
        
        # COGS Items
        cogs_items = analysis_data.get('cost_of_goods_sold', [])
        if cogs_items:
            ws[f'A{row}'] = "Cost of Goods Sold (COGS)"
            ws[f'A{row}'].font = self.title_font
            row += 1
            
            ws[f'A{row}'] = "Product Category"
            ws[f'B{row}'] = "Price/Item (€)"
            ws[f'C{row}'] = "COGS/Item (€)"
            ws[f'D{row}'] = "Gross Margin %"
            ws[f'E{row}'] = "Reasoning"
            self.style_header_row(ws, row, 5)
            row += 1
            
            for item in cogs_items:
                ws[f'A{row}'] = item.get('product_category', 'N/A')
                ws[f'B{row}'] = item.get('price_per_item', 0)
                ws[f'B{row}'].number_format = '€#,##0.00'
                ws[f'C{row}'] = item.get('cogs_per_item', 0)
                ws[f'C{row}'].number_format = '€#,##0.00'
                ws[f'D{row}'] = item.get('gross_margin_percentage', 0) / 100
                ws[f'D{row}'].number_format = '0.0%'
                ws[f'E{row}'] = item.get('reasoning', 'N/A')
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].border = self.border
                row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 45
    
    def create_financial_projections_sheet(self, wb: Workbook, analysis_data: Dict[str, Any]):
        """Create Financial Projections sheet"""
        ws = wb.create_sheet("Financial Projections")
        row = 1
        
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "7-Year Financial Projections"
        ws[f'A{row}'].font = Font(bold=True, size=16, color="1F4788")
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        yearly_breakdown = analysis_data.get('yearly_cost_breakdown', {})
        if yearly_breakdown:
            ws[f'A{row}'] = "Year"
            ws[f'B{row}'] = "Volume"
            ws[f'C{row}'] = "Development (€)"
            ws[f'D{row}'] = "CAC (€)"
            ws[f'E{row}'] = "Operations (€)"
            ws[f'F{row}'] = "COGS (€)"
            ws[f'G{row}'] = "Total Cost (€)"
            ws[f'H{row}'] = "Cost/Unit (€)"
            self.style_header_row(ws, row, 8)
            row += 1
            
            for year, data in sorted(yearly_breakdown.items()):
                ws[f'A{row}'] = year
                ws[f'B{row}'] = data.get('projected_volume', 0)
                ws[f'B{row}'].number_format = '#,##0'
                ws[f'C{row}'] = data.get('one_time_development', 0)
                ws[f'C{row}'].number_format = '€#,##0.00'
                ws[f'D{row}'] = data.get('customer_acquisition', 0)
                ws[f'D{row}'].number_format = '€#,##0.00'
                ws[f'E{row}'] = data.get('distribution_operations', 0)
                ws[f'E{row}'].number_format = '€#,##0.00'
                ws[f'F{row}'] = data.get('total_cogs', 0)
                ws[f'F{row}'].number_format = '€#,##0.00'
                ws[f'G{row}'] = data.get('total_cost', 0)
                ws[f'G{row}'].number_format = '€#,##0.00'
                ws[f'H{row}'] = data.get('cogs_per_unit', 0)
                ws[f'H{row}'].number_format = '€#,##0.00'
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = self.border
                row += 1
            
            # 7-Year Summary
            seven_year = analysis_data.get('seven_year_summary', {})
            if seven_year:
                row += 1
                ws[f'A{row}'] = "7-Year Summary"
                ws[f'A{row}'].font = self.title_font
                row += 1
                
                ws[f'A{row}'] = "Total Cost (2024-2030):"
                ws[f'B{row}'] = seven_year.get('total_cost_2024_2030', 0)
                ws[f'B{row}'].number_format = '€#,##0.00'
                ws[f'B{row}'].font = self.currency_font
                row += 1
                
                ws[f'A{row}'] = "Total Volume (2024-2030):"
                ws[f'B{row}'] = seven_year.get('total_volume_2024_2030', 0)
                ws[f'B{row}'].number_format = '#,##0'
                row += 1
                
                ws[f'A{row}'] = "Average Cost per Unit:"
                ws[f'B{row}'] = seven_year.get('average_cost_per_unit', 0)
                ws[f'B{row}'].number_format = '€#,##0.00'
                ws[f'B{row}'].font = self.currency_font
        
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 16
    
    def create_risks_strategy_sheet(self, wb: Workbook, analysis_data: Dict[str, Any]):
        """Create Risks & Strategy sheet"""
        ws = wb.create_sheet("Risks & Strategy")
        row = 1
        
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "Risk Assessment & Competitive Strategy"
        ws[f'A{row}'].font = Font(bold=True, size=16, color="1F4788")
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        # Key Risks
        key_risks = analysis_data.get('key_risks', [])
        if key_risks:
            ws[f'A{row}'] = "Key Risks"
            ws[f'A{row}'].font = self.title_font
            row += 1
            
            ws[f'A{row}'] = "Risk"
            ws[f'B{row}'] = "Probability"
            ws[f'C{row}'] = "Impact"
            ws[f'D{row}'] = "Mitigation Strategy"
            self.style_header_row(ws, row, 4)
            row += 1
            
            for risk in key_risks:
                ws[f'A{row}'] = risk.get('risk', 'N/A')
                ws[f'B{row}'] = risk.get('probability', 'N/A')
                ws[f'C{row}'] = risk.get('impact', 'N/A')
                ws[f'D{row}'] = risk.get('mitigation', 'N/A')
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].border = self.border
                    ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                row += 1
            row += 1
        
        # Competitive Advantages
        comp_adv = analysis_data.get('competitive_advantages', [])
        if comp_adv:
            ws[f'A{row}'] = "Competitive Advantages"
            ws[f'A{row}'].font = self.title_font
            row += 1
            
            ws[f'A{row}'] = "Advantage"
            ws[f'B{row}'] = "Market Validation"
            ws[f'C{row}'] = "Sustainability"
            self.style_header_row(ws, row, 3)
            row += 1
            
            for adv in comp_adv:
                ws[f'A{row}'] = adv.get('advantage', 'N/A')
                ws[f'B{row}'] = adv.get('market_validation', 'N/A')
                ws[f'C{row}'] = adv.get('sustainability_assessment', 'N/A')
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = self.border
                    ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                row += 1
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40
    
    def create_complete_metrics_sheet(self, wb: Workbook, analysis_data: Dict[str, Any]):
        """Create comprehensive dashboard metrics sheet with ALL financial data"""
        ws = wb.create_sheet("Complete Financial Analysis")
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Complete Financial Analysis & Metrics Dashboard"
        ws[f'A{row}'].font = Font(bold=True, size=16, color="1F4788")
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        # ====================
        # KEY FINANCIAL METRICS
        # ====================
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "KEY FINANCIAL METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Headers
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'C{row}'] = "Confidence"
        ws[f'D{row}'] = "Details"
        ws[f'E{row}'] = "Year 1"
        ws[f'F{row}'] = "5-Year Total"
        self.style_subheader_row(ws, row, 6)
        row += 1
        
        # TAM
        tam = analysis_data.get('tam', {})
        ws[f'A{row}'] = "TAM (Total Addressable Market)"
        ws[f'B{row}'] = tam.get('market_size', 0)
        ws[f'B{row}'].number_format = '€#,##0'
        ws[f'C{row}'] = f"{tam.get('confidence', 0)}%"
        ws[f'D{row}'] = f"Market size with {tam.get('market_growth_rate', 0):.1f}% annual growth"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # SAM
        sam = analysis_data.get('sam', {})
        ws[f'A{row}'] = "SAM (Serviceable Addressable Market)"
        ws[f'B{row}'] = sam.get('market_size', 0)
        ws[f'B{row}'].number_format = '€#,##0'
        ws[f'C{row}'] = f"{sam.get('confidence', 0)}%"
        ws[f'D{row}'] = "Realistic serviceable market"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # SOM
        som = analysis_data.get('som', {})
        ws[f'A{row}'] = "SOM (Serviceable Obtainable Market)"
        ws[f'B{row}'] = som.get('revenue_potential', 0)
        ws[f'B{row}'].number_format = '€#,##0'
        ws[f'C{row}'] = f"{som.get('confidence', 0)}%"
        ws[f'D{row}'] = "Achievable revenue potential"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # ROI
        roi = analysis_data.get('roi', {})
        ws[f'A{row}'] = "ROI (Return on Investment)"
        ws[f'B{row}'] = roi.get('roi_percentage', 0) / 100
        ws[f'B{row}'].number_format = '0.0%'
        ws[f'C{row}'] = f"{roi.get('confidence', 0)}%"
        ws[f'D{row}'] = f"Break-even in {roi.get('payback_period_months', 0)} months"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # Turnover
        turnover = analysis_data.get('turnover', {})
        turnover_numbers = turnover.get('numbers', {}) if turnover else {}
        year1_revenue = turnover_numbers.get('2024', 0) if turnover_numbers else 0
        total_revenue_5y = sum(list(turnover_numbers.values())[:5]) if turnover_numbers else 0
        
        ws[f'A{row}'] = "Turnover (Revenue)"
        ws[f'B{row}'] = turnover.get('total_revenue', year1_revenue) if turnover else 0
        ws[f'B{row}'].number_format = '€#,##0'
        ws[f'C{row}'] = f"{turnover.get('confidence', 0) if turnover else 0}%"
        ws[f'D{row}'] = f"Average annual revenue with {turnover.get('yoy_growth', 0)*100 if turnover else 0:.1f}% YoY growth"
        ws[f'E{row}'] = year1_revenue
        ws[f'E{row}'].number_format = '€#,##0'
        ws[f'F{row}'] = total_revenue_5y
        ws[f'F{row}'].number_format = '€#,##0'
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # Volume
        volume = analysis_data.get('volume', {})
        ws[f'A{row}'] = "Volume (Units Sold)"
        ws[f'B{row}'] = volume.get('units_sold', 0)
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'] = f"{volume.get('confidence', 0)}%"
        ws[f'D{row}'] = f"Projected units in year 1"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # Unit Economics
        unit_econ = analysis_data.get('unit_economics', {})
        ws[f'A{row}'] = "Unit Economics (Profit per Unit)"
        ws[f'B{row}'] = unit_econ.get('profit_per_unit', 0)
        ws[f'B{row}'].number_format = '€#,##0.00'
        ws[f'C{row}'] = f"{unit_econ.get('confidence', 0)}%"
        ws[f'D{row}'] = f"Profit margin: {unit_econ.get('margin_percentage', 0):.1f}%"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # EBIT
        ebit = analysis_data.get('ebit', {})
        ebit_numbers = ebit.get('numbers', {}) if ebit else {}
        year1_ebit = ebit_numbers.get('2024', 0) if ebit_numbers else 0
        total_ebit_5y = sum(list(ebit_numbers.values())[:5]) if ebit_numbers else 0
        
        ws[f'A{row}'] = "EBIT (Earnings Before Interest & Tax)"
        ws[f'B{row}'] = ebit.get('ebit_value', year1_ebit) if ebit else 0
        ws[f'B{row}'].number_format = '€#,##0'
        ws[f'C{row}'] = f"{ebit.get('confidence', 0) if ebit else 0}%"
        ws[f'D{row}'] = f"EBIT margin: {ebit.get('ebit_percentage', 0) if ebit else 0:.1f}%"
        ws[f'E{row}'] = year1_ebit
        ws[f'E{row}'].number_format = '€#,##0'
        ws[f'F{row}'] = total_ebit_5y
        ws[f'F{row}'].number_format = '€#,##0'
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # COGS
        cogs = analysis_data.get('cogs', {})
        ws[f'A{row}'] = "COGS (Cost of Goods Sold)"
        ws[f'B{row}'] = cogs.get('total_cogs', 0)
        ws[f'B{row}'].number_format = '€#,##0'
        ws[f'C{row}'] = f"{cogs.get('confidence', 0)}%"
        ws[f'D{row}'] = f"COGS per unit: €{cogs.get('cogs_per_unit', 0):.2f}"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # Market Potential
        market_potential = analysis_data.get('market_potential', {})
        ws[f'A{row}'] = "Market Potential"
        ws[f'B{row}'] = market_potential.get('market_size', 0)
        ws[f'B{row}'].number_format = '€#,##0'
        ws[f'C{row}'] = f"{market_potential.get('confidence', 0)}%"
        ws[f'D{row}'] = f"Strong market potential with {market_potential.get('growth_rate', 0):.1f}% growth"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = self.border
        row += 2
        
        # ====================
        # PROFIT & LOSS SUMMARY
        # ====================
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "PROFIT & LOSS SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        ws[f'A{row}'] = "Item"
        ws[f'B{row}'] = "Amount (€)"
        ws[f'C{row}'] = "% of Revenue"
        ws[f'D{row}'] = "Description"
        self.style_subheader_row(ws, row, 4)
        ws.merge_cells(f'D{row}:F{row}')
        row += 1
        
        # Revenue
        ws[f'A{row}'] = "Total Revenue"
        ws[f'B{row}'] = year1_revenue
        ws[f'B{row}'].number_format = '€#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color="006400")
        ws[f'C{row}'] = 1.0
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'D{row}'] = "Gross income from all sales"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # COGS Breakdown
        ws[f'A{row}'] = "Cost of Goods Sold (COGS)"
        ws[f'B{row}'] = cogs.get('total_cogs', 0)
        ws[f'B{row}'].number_format = '€#,##0.00'
        ws[f'B{row}'].font = Font(color="8B0000")
        cogs_pct = (cogs.get('total_cogs', 0) / year1_revenue) if year1_revenue > 0 else 0
        ws[f'C{row}'] = cogs_pct
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'D{row}'] = "Direct costs of producing goods"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # Material Costs
        ws[f'A{row}'] = "  - Material Costs"
        ws[f'B{row}'] = cogs.get('material', 0)
        ws[f'B{row}'].number_format = '€#,##0.00'
        mat_pct = (cogs.get('material', 0) / year1_revenue) if year1_revenue > 0 else 0
        ws[f'C{row}'] = mat_pct
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'D{row}'] = "Raw materials and components"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # Labor Costs
        ws[f'A{row}'] = "  - Labor Costs"
        ws[f'B{row}'] = cogs.get('labor', 0)
        ws[f'B{row}'].number_format = '€#,##0.00'
        labor_pct = (cogs.get('labor', 0) / year1_revenue) if year1_revenue > 0 else 0
        ws[f'C{row}'] = labor_pct
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'D{row}'] = "Direct labor costs"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # Overhead Costs
        ws[f'A{row}'] = "  - Overhead Costs"
        ws[f'B{row}'] = cogs.get('overheads', 0)
        ws[f'B{row}'].number_format = '€#,##0.00'
        overhead_pct = (cogs.get('overheads', 0) / year1_revenue) if year1_revenue > 0 else 0
        ws[f'C{row}'] = overhead_pct
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'D{row}'] = "Manufacturing overhead"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # Gross Profit
        gross_profit = year1_revenue - cogs.get('total_cogs', 0)
        ws[f'A{row}'] = "Gross Profit"
        ws[f'B{row}'] = gross_profit
        ws[f'B{row}'].number_format = '€#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color="006400")
        gross_margin = (gross_profit / year1_revenue) if year1_revenue > 0 else 0
        ws[f'C{row}'] = gross_margin
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'D{row}'] = "Revenue minus COGS"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = self.border
            ws[f'{col}{row}'].fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        row += 1
        
        # Operating Expenses
        roi_costs = roi.get('cost_breakdown', {}) if roi.get('cost_breakdown') else {}
        year1_cost = roi_costs.get('2024', 0) if roi_costs else 0
        operating_expenses = year1_cost - cogs.get('total_cogs', 0)
        
        ws[f'A{row}'] = "Operating Expenses"
        ws[f'B{row}'] = operating_expenses
        ws[f'B{row}'].number_format = '€#,##0.00'
        ws[f'B{row}'].font = Font(color="8B0000")
        opex_pct = (operating_expenses / year1_revenue) if year1_revenue > 0 else 0
        ws[f'C{row}'] = opex_pct
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'D{row}'] = "Marketing, R&D, admin costs"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = self.border
        row += 1
        
        # EBIT / Operating Profit
        ws[f'A{row}'] = "EBIT (Operating Profit)"
        ws[f'B{row}'] = year1_ebit
        ws[f'B{row}'].number_format = '€#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color="006400")
        ebit_margin = (year1_ebit / year1_revenue) if year1_revenue > 0 else 0
        ws[f'C{row}'] = ebit_margin
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'D{row}'] = "Earnings before interest & tax"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = self.border
            ws[f'{col}{row}'].fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        row += 1
        
        # Net Profit (assuming EBIT = Net Profit for simplicity)
        net_profit = year1_revenue - year1_cost
        ws[f'A{row}'] = "Net Profit"
        ws[f'B{row}'] = net_profit
        ws[f'B{row}'].number_format = '€#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, size=12, color="006400")
        net_margin = (net_profit / year1_revenue) if year1_revenue > 0 else 0
        ws[f'C{row}'] = net_margin
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'D{row}'] = "Bottom line profit"
        ws.merge_cells(f'D{row}:F{row}')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = self.border
            ws[f'{col}{row}'].fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        row += 3
        row += 3
        
        # ====================
        # 5-YEAR INCOME PROJECTIONS
        # ====================
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "5-YEAR INCOME PROJECTIONS"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Headers for projection table
        ws[f'A{row}'] = "Year"
        ws[f'B{row}'] = "Revenue (€)"
        ws[f'C{row}'] = "COGS (€)"
        ws[f'D{row}'] = "Gross Profit (€)"
        ws[f'E{row}'] = "Operating Costs (€)"
        ws[f'F{row}'] = "EBIT (€)"
        ws[f'G{row}'] = "Net Profit (€)"
        ws[f'H{row}'] = "Profit Margin %"
        self.style_header_row(ws, row, 8)
        row += 1
        
        # Get yearly data
        years_to_show = ['2024', '2025', '2026', '2027', '2028']
        
        total_5y_revenue = 0
        total_5y_cogs_val = 0
        total_5y_gross = 0
        total_5y_opex = 0
        total_5y_ebit_val = 0
        total_5y_net = 0
        
        for year in years_to_show:
            year_revenue = turnover_numbers.get(year, 0) if turnover_numbers else 0
            year_cost = roi_costs.get(year, 0) if roi_costs else 0
            
            # Estimate COGS as percentage of revenue
            cogs_pct_val = cogs.get('cogs_percentage', 85) / 100 if cogs else 0.85
            year_cogs = year_revenue * cogs_pct_val
            year_gross = year_revenue - year_cogs
            year_opex = year_cost - year_cogs if year_cost > year_cogs else 0
            year_ebit_val = ebit_numbers.get(year, year_gross - year_opex) if ebit_numbers else (year_gross - year_opex)
            year_net = year_revenue - year_cost
            year_margin = (year_net / year_revenue * 100) if year_revenue > 0 else 0
            
            ws[f'A{row}'] = year
            ws[f'B{row}'] = year_revenue
            ws[f'B{row}'].number_format = '€#,##0'
            ws[f'C{row}'] = year_cogs
            ws[f'C{row}'].number_format = '€#,##0'
            ws[f'D{row}'] = year_gross
            ws[f'D{row}'].number_format = '€#,##0'
            ws[f'E{row}'] = year_opex
            ws[f'E{row}'].number_format = '€#,##0'
            ws[f'F{row}'] = year_ebit_val
            ws[f'F{row}'].number_format = '€#,##0'
            ws[f'G{row}'] = year_net
            ws[f'G{row}'].number_format = '€#,##0'
            ws[f'H{row}'] = year_margin / 100
            ws[f'H{row}'].number_format = '0.0%'
            
            total_5y_revenue += year_revenue
            total_5y_cogs_val += year_cogs
            total_5y_gross += year_gross
            total_5y_opex += year_opex
            total_5y_ebit_val += year_ebit_val
            total_5y_net += year_net
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        # Totals row
        ws[f'A{row}'] = "5-Year Total"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'B{row}'] = total_5y_revenue
        ws[f'B{row}'].number_format = '€#,##0'
        ws[f'B{row}'].font = self.currency_font
        ws[f'C{row}'] = total_5y_cogs_val
        ws[f'C{row}'].number_format = '€#,##0'
        ws[f'C{row}'].font = self.currency_font
        ws[f'D{row}'] = total_5y_gross
        ws[f'D{row}'].number_format = '€#,##0'
        ws[f'D{row}'].font = self.currency_font
        ws[f'E{row}'] = total_5y_opex
        ws[f'E{row}'].number_format = '€#,##0'
        ws[f'E{row}'].font = self.currency_font
        ws[f'F{row}'] = total_5y_ebit_val
        ws[f'F{row}'].number_format = '€#,##0'
        ws[f'F{row}'].font = self.currency_font
        ws[f'G{row}'] = total_5y_net
        ws[f'G{row}'].number_format = '€#,##0'
        ws[f'G{row}'].font = Font(bold=True, size=11, color="006400")
        avg_margin = (total_5y_net / total_5y_revenue * 100) if total_5y_revenue > 0 else 0
        ws[f'H{row}'] = avg_margin / 100
        ws[f'H{row}'].number_format = '0.0%'
        ws[f'H{row}'].font = Font(bold=True)
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = self.border
            ws.cell(row=row, column=col).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        row += 3
        
        # ====================
        # BUSINESS INSIGHTS
        # ====================
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "BUSINESS INSIGHTS & RECOMMENDATIONS"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Key Insights
        ws[f'A{row}'] = "Key Financial Insights"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4788")
        row += 1
        
        insights = [
            f"• Volume Projection: {volume.get('insight', 'N/A')}",
            f"• Revenue Growth: {turnover.get('insight', 'N/A')}",
            f"• Profitability: {unit_econ.get('insight', 'N/A')}",
            f"• Market Position: {market_potential.get('insight', 'N/A')}",
            f"• EBIT Performance: {ebit.get('insight', 'N/A')}",
            f"• ROI Analysis: {roi.get('insight', 'N/A')}"
        ]
        
        for insight in insights:
            ws[f'A{row}'] = insight
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 20
            row += 1
        
        row += 1
        
        # Business Assumptions
        assumptions = analysis_data.get('business_assumptions', [])
        if assumptions:
            ws[f'A{row}'] = "Business Assumptions"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4788")
            row += 1
            
            for i, assumption in enumerate(assumptions[:8], 1):  # Show top 8
                ws[f'A{row}'] = f"{i}. {assumption}"
                ws.merge_cells(f'A{row}:F{row}')
                ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                ws.row_dimensions[row].height = 18
                row += 1
            row += 1
        
        # Improvement Recommendations
        recommendations = analysis_data.get('improvement_recommendations', [])
        if recommendations:
            ws[f'A{row}'] = "Strategic Recommendations"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4788")
            row += 1
            
            for i, rec in enumerate(recommendations[:8], 1):  # Show top 8
                ws[f'A{row}'] = f"{i}. {rec}"
                ws.merge_cells(f'A{row}:F{row}')
                ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                ws.row_dimensions[row].height = 18
                row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 15
    
    def generate_excel(self, analysis_data: Dict[str, Any]) -> str:
        """
        Generate complete Excel workbook with all sheets
        Returns: filepath to generated Excel file
        """
        print("\n" + "="*80)
        print("📊 EXCEL EXPORT: Creating professional comprehensive report")
        print("="*80)
        
        wb = Workbook()
        
        # Create all sheets in order
        self.create_executive_summary_sheet(wb, analysis_data)
        self.create_complete_metrics_sheet(wb, analysis_data)  # Complete financial analysis with P&L
        self.create_cost_breakdown_sheet(wb, analysis_data)
        self.create_financial_projections_sheet(wb, analysis_data)
        self.create_risks_strategy_sheet(wb, analysis_data)
        
        # Save the file
        filename = f"BMW_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join("exports", filename)
        wb.save(filepath)
        
        print(f"✅ Excel file created: {filepath}")
        print(f"📦 File size: {os.path.getsize(filepath) / 1024:.2f} KB")
        print(f"📋 Sheets: Executive Summary, Complete Financial Analysis, Cost Breakdown, Projections, Risks & Strategy")
        print("="*80 + "\n")
        
        return filepath
